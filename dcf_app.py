"""
Universal DCF Valuation Model — Any Public Company
Built by: Poonam Dhanuka | DePaul MS Finance 2027
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io
from datetime import datetime

# ── PAGE CONFIG ──────────────────────────────────────────────────
st.set_page_config(
    page_title="DCF Valuation Model",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: #0f0f0f; color: #e0e0e0; }
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
section[data-testid="stSidebar"] { background-color: #111 !important; border-right: 1px solid #222 !important; }
section[data-testid="stSidebar"] * { color: #ccc !important; }
.dcf-header { background: linear-gradient(135deg, #1B3A6B 0%, #2d5aa0 100%); border-radius: 14px; padding: 24px 28px; margin-bottom: 20px; }
.dcf-title  { font-size: 26px; font-weight: 800; color: #fff; }
.dcf-sub    { font-size: 13px; color: #aac; margin-top: 4px; }
.company-pill { display:inline-block; background:#1B3A6B; color:#fff; border-radius:20px; padding:4px 14px; font-size:13px; font-weight:700; margin-right:8px; }
.result-card  { background:#141414; border:1px solid #222; border-radius:12px; padding:20px 24px; text-align:center; }
.result-label { font-size:11px; color:#666; text-transform:uppercase; letter-spacing:0.8px; }
.result-value { font-size:32px; font-weight:800; color:#fff; margin-top:4px; }
.result-sub   { font-size:12px; color:#888; margin-top:2px; }
.metric-card  { background:#141414; border:1px solid #222; border-radius:10px; padding:14px 16px; margin-bottom:8px; }
.metric-label { font-size:10px; color:#666; text-transform:uppercase; letter-spacing:0.5px; }
.metric-value { font-size:18px; font-weight:700; color:#fff; margin-top:3px; }
.section-title { font-size:15px; font-weight:700; color:#fff; border-left:3px solid #1B3A6B; padding-left:10px; margin:20px 0 12px 0; }
.info-box { background:#1a1a1a; border:1px solid #2a2a2a; border-radius:8px; padding:10px 14px; font-size:12px; color:#888; margin:8px 0; }
hr { border-color: #222 !important; }
[data-testid="metric-container"] { background:#141414; border:1px solid #222; border-radius:10px; padding:14px 16px !important; }
[data-testid="metric-container"] label { color:#888 !important; font-size:11px !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color:#fff !important; font-weight:700 !important; }
.stTabs [data-baseweb="tab-list"] { background:#141414; border-radius:8px; }
.stTabs [data-baseweb="tab"]      { color:#888 !important; }
.stTabs [aria-selected="true"]    { background:#1B3A6B !important; color:#fff !important; }
.stButton > button { background:#1B3A6B !important; color:#fff !important; border:none !important; border-radius:8px !important; font-weight:600 !important; }
.stTextInput input { background:#1a1a1a !important; color:#fff !important; border-color:#333 !important; border-radius:8px !important; font-size:16px !important; font-weight:700 !important; }
.stNumberInput input { background:#1a1a1a !important; color:#fff !important; border-color:#333 !important; border-radius:8px !important; }
</style>
""", unsafe_allow_html=True)

DARK = dict(
    paper_bgcolor='#111', plot_bgcolor='#111',
    font=dict(color='#aaa', size=12),
    xaxis=dict(showgrid=False, color='#444', zeroline=False),
    yaxis=dict(gridcolor='#222', color='#aaa', zeroline=False),
    margin=dict(l=10, r=10, t=40, b=10),
    legend=dict(bgcolor='rgba(0,0,0,0)', font=dict(color='#aaa'))
)

YEARS = [2026, 2027, 2028, 2029, 2030]

# ── COMPANY NAME → TICKER LOOKUP ─────────────────────────────────
COMPANY_LOOKUP = {
    "Apple": "AAPL", "Microsoft": "MSFT", "Google": "GOOGL", "Alphabet": "GOOGL",
    "Amazon": "AMZN", "Tesla": "TSLA", "Meta": "META", "Facebook": "META",
    "Nvidia": "NVDA", "Netflix": "NFLX", "PepsiCo": "PEP", "Pepsi": "PEP",
    "Coca-Cola": "KO", "Coke": "KO", "Johnson & Johnson": "JNJ", "J&J": "JNJ",
    "JPMorgan": "JPM", "JP Morgan": "JPM", "Goldman Sachs": "GS", "Goldman": "GS",
    "Walmart": "WMT", "Disney": "DIS", "Walt Disney": "DIS",
    "Nike": "NKE", "Starbucks": "SBUX", "McDonald's": "MCD", "McDonalds": "MCD",
    "Visa": "V", "Mastercard": "MA", "PayPal": "PYPL",
    "Bank of America": "BAC", "Wells Fargo": "WFC", "Citigroup": "C", "Citi": "C",
    "ExxonMobil": "XOM", "Exxon": "XOM", "Chevron": "CVX",
    "Procter & Gamble": "PG", "P&G": "PG", "Unilever": "UL",
    "Intel": "INTC", "AMD": "AMD", "Qualcomm": "QCOM", "Broadcom": "AVGO",
    "Salesforce": "CRM", "Oracle": "ORCL", "SAP": "SAP", "IBM": "IBM",
    "Boeing": "BA", "Caterpillar": "CAT", "3M": "MMM",
    "Pfizer": "PFE", "Moderna": "MRNA", "AbbVie": "ABBV", "Merck": "MRK",
    "UnitedHealth": "UNH", "CVS": "CVS", "Costco": "COST", "Target": "TGT",
    "Home Depot": "HD", "Lowe's": "LOW", "Lowes": "LOW",
    "Uber": "UBER", "Lyft": "LYFT", "Airbnb": "ABNB", "Booking": "BKNG",
    "Spotify": "SPOT", "Snap": "SNAP", "Twitter": "X", "X": "X",
    "Adobe": "ADBE", "Shopify": "SHOP", "Zoom": "ZM", "Slack": "CRM",
    "Goldman": "GS", "Morgan Stanley": "MS", "BlackRock": "BLK",
    "AT&T": "T", "Verizon": "VZ", "Comcast": "CMCSA",
    "Berkshire": "BRK-B", "Berkshire Hathaway": "BRK-B",
    "TSMC": "TSM", "Samsung": "SSNLF", "Sony": "SONY",
    "Kellogg": "K", "General Mills": "GIS", "Kraft Heinz": "KHC",
    "Colgate": "CL", "Kimberly-Clark": "KMB",
    "FedEx": "FDX", "UPS": "UPS", "DHL": "DHL",
}

@st.cache_data(ttl=300)
def search_ticker(query):
    """Search Yahoo Finance for company name → return list of (name, ticker) matches."""
    query = query.strip()
    if not query or len(query) < 2:
        return []
    try:
        results = yf.Search(query, max_results=6)
        quotes  = results.quotes if hasattr(results, "quotes") else []
        out = []
        for q in quotes:
            sym  = q.get("symbol", "")
            name = q.get("longname") or q.get("shortname") or sym
            qtype = q.get("quoteType", "")
            if sym and qtype in ("EQUITY", "ETF"):
                out.append((name, sym))
        return out
    except:
        return []

# ─────────────────────────────────────────────────────────────────
# DATA FETCHER
# ─────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def fetch_company_data(ticker_sym):
    """Fetch real financials from Yahoo Finance for any ticker."""
    try:
        t    = yf.Ticker(ticker_sym)
        info = t.info

        # ── Basic info ──────────────────────────────────────────
        name     = info.get("longName") or info.get("shortName") or ticker_sym
        sector   = info.get("sector", "N/A")
        industry = info.get("industry", "N/A")
        currency = info.get("financialCurrency", "USD")
        exchange = info.get("exchange", "")

        # ── Key financials (in units as reported by yfinance) ───
        revenue    = info.get("totalRevenue", 0) or 0
        ebitda     = info.get("ebitda", 0) or 0
        ebit       = info.get("ebit", 0) or info.get("operatingIncome", 0) or 0
        net_income = info.get("netIncomeToCommon", 0) or 0
        total_debt = info.get("totalDebt", 0) or 0
        cash       = info.get("totalCash", 0) or 0
        shares     = info.get("sharesOutstanding", 0) or 0
        beta       = info.get("beta") or 0.85
        price      = info.get("currentPrice") or info.get("regularMarketPrice") or 0
        mktcap     = info.get("marketCap", 0) or 0

        # Try getting operating cash flow and capex from cashflow statement
        try:
            cf = t.cashflow
            if cf is not None and not cf.empty:
                # Most recent year
                op_cf  = float(cf.loc["Operating Cash Flow"].iloc[0])  if "Operating Cash Flow" in cf.index else 0
                capex  = abs(float(cf.loc["Capital Expenditure"].iloc[0])) if "Capital Expenditure" in cf.index else revenue * 0.05
                da     = float(cf.loc["Depreciation And Amortization"].iloc[0]) if "Depreciation And Amortization" in cf.index else revenue * 0.04
            else:
                capex = revenue * 0.05
                da    = revenue * 0.04
        except:
            capex = revenue * 0.05
            da    = revenue * 0.04

        # Get tax rate from income statement
        try:
            fin = t.financials
            if fin is not None and not fin.empty:
                tax_exp  = abs(float(fin.loc["Tax Provision"].iloc[0])) if "Tax Provision" in fin.index else 0
                pretax   = float(fin.loc["Pretax Income"].iloc[0]) if "Pretax Income" in fin.index else ebit
                tax_rate = round(tax_exp / pretax, 3) if pretax > 0 else 0.21
                tax_rate = max(0.05, min(tax_rate, 0.40))  # clamp 5%-40%
            else:
                tax_rate = 0.21
        except:
            tax_rate = 0.21

        # Convert to millions
        def to_mn(v): return round(v / 1e6) if v else 0

        return {
            "ok": True,
            "name": name, "ticker": ticker_sym.upper(),
            "sector": sector, "industry": industry,
            "currency": currency, "exchange": exchange,
            "price": price, "mktcap": to_mn(mktcap),
            "revenue": to_mn(revenue),
            "ebit":    to_mn(ebit),
            "ebitda":  to_mn(ebitda),
            "da":      to_mn(da),
            "capex":   to_mn(capex),
            "debt":    to_mn(total_debt),
            "cash":    to_mn(cash),
            "shares":  round(shares / 1e6, 1),
            "beta":    round(beta, 2),
            "tax_rate": round(tax_rate, 3),
            "ebit_margin": round(ebit / revenue, 4) if revenue else 0.12,
            "da_pct":   round(da / revenue, 4) if revenue else 0.04,
            "capex_pct":round(capex / revenue, 4) if revenue else 0.05,
            "non_op":  round(to_mn(cash) * 0.1),  # rough proxy
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 🔍 Company Search")

    # Show currently loaded company
    current_ticker = st.session_state.get("ticker", "PEP")
    st.markdown(f"""
    <div style="background:#1B3A6B; border-radius:8px; padding:8px 12px; margin-bottom:10px;
                font-size:12px; color:#aac;">
        Currently loaded: <b style="color:#fff; font-size:14px">{current_ticker}</b>
        &nbsp;·&nbsp; <span style="color:#aac">Type below to change</span>
    </div>
    """, unsafe_allow_html=True)

    search_query = st.text_input(
        "Search Company",
        value="",
        placeholder="e.g. Apple, Tesla, PepsiCo, MSFT...",
        key="search_input"
    ).strip()

    # ── Smart lookup ──────────────────────────────────────────────
    resolved_ticker = None

    if search_query:
        # 1. Check local dictionary first (instant, no API call)
        for name, sym in COMPANY_LOOKUP.items():
            if search_query.lower() in name.lower():
                resolved_ticker = sym
                break

        # 2. If looks like a ticker (short, uppercase-able) → use directly
        if not resolved_ticker and len(search_query) <= 5 and search_query.replace("-","").isalpha():
            resolved_ticker = search_query.upper()

        # 3. Live search via yfinance — store results in session so dropdown persists
        if not resolved_ticker:
            with st.spinner("Searching..."):
                results = search_ticker(search_query)
            if results:
                st.session_state["last_search_results"] = results
                st.session_state["last_search_query"]   = search_query
            else:
                st.warning("No results found. Try a different name or ticker.")

        # Show dropdown if we have live results for this query
        if not resolved_ticker and st.session_state.get("last_search_query") == search_query:
            results = st.session_state.get("last_search_results", [])
            if results:
                options = [f"{name}  ({sym})" for name, sym in results]
                chosen  = st.selectbox("Select company", options,
                                       key="company_select", label_visibility="collapsed")
                resolved_ticker = chosen.split("(")[-1].rstrip(")")

        if resolved_ticker:
            st.success(f"✅ Selected: **{resolved_ticker}**")

    # Popular quick-pick buttons
    st.markdown("**⚡ Quick Pick**")
    popular = [("🍎 Apple","AAPL"),("🪟 Microsoft","MSFT"),("🔍 Google","GOOGL"),
               ("🛒 Amazon","AMZN"),("🚗 Tesla","TSLA"),("🟢 Nvidia","NVDA"),
               ("👟 Nike","NKE"),("🥤 PepsiCo","PEP"),("💳 Visa","V"),("📱 Meta","META")]
    cols = st.columns(2)
    for i, (label, sym) in enumerate(popular):
        with cols[i % 2]:
            if st.button(label, key=f"quick_{sym}", use_container_width=True):
                resolved_ticker = sym
                st.session_state.ticker = sym

    if resolved_ticker:
        st.session_state.ticker = resolved_ticker

    final_ticker = st.session_state.get("ticker", "PEP")
    st.caption("Data auto-fetched live from Yahoo Finance")
    st.divider()

    st.markdown("**📈 Revenue Growth (Override)**")
    g1 = st.slider("2026 %", 0.0, 20.0, 4.0, 0.5) / 100
    g2 = st.slider("2027 %", 0.0, 20.0, 4.0, 0.5) / 100
    g3 = st.slider("2028 %", 0.0, 20.0, 3.5, 0.5) / 100
    g4 = st.slider("2029 %", 0.0, 20.0, 3.5, 0.5) / 100
    g5 = st.slider("2030 %", 0.0, 20.0, 3.0, 0.5) / 100
    revenue_growth = [g1, g2, g3, g4, g5]

    st.divider()
    st.markdown("**🏦 WACC Inputs**")
    rf_rate  = st.slider("Risk-Free Rate %",       2.0, 7.0, 4.5, 0.25) / 100
    erp      = st.slider("Equity Risk Premium %",  3.0, 8.0, 5.5, 0.25) / 100
    add_rp   = st.slider("Additional Risk Premium %", 0.0, 3.0, 0.0, 0.25) / 100
    pretax_kd= st.slider("Pre-Tax Cost of Debt %", 2.0, 8.0, 3.5, 0.25) / 100
    debt_wt  = st.slider("Debt Weight %",         10.0, 80.0, 40.0, 5.0) / 100

    st.divider()
    st.markdown("**🎯 Terminal Value**")
    ltgr     = st.slider("Long-Term Growth Rate %", 1.0, 4.0, 2.5, 0.25) / 100
    exit_mult= st.slider("Exit EV/EBITDA Multiple", 6.0, 25.0, 12.0, 0.5)

    st.divider()
    st.caption("Built by Poonam Dhanuka · DePaul MS Finance 2027")


# ─────────────────────────────────────────────────────────────────
# FETCH DATA
# ─────────────────────────────────────────────────────────────────

if "ticker" not in st.session_state:
    st.session_state.ticker = "PEP"

ticker_sym = st.session_state.get("ticker", "PEP")

with st.spinner(f"Fetching data for {ticker_sym} from Yahoo Finance..."):
    d = fetch_company_data(ticker_sym)

if not d["ok"]:
    st.error(f"❌ Could not load data for **{ticker_sym}**. Check the ticker symbol and try again.")
    st.stop()

# ─────────────────────────────────────────────────────────────────
# OVERRIDE SLIDERS (auto-populated from real data, user can adjust)
# ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.divider()
    st.markdown(f"**⚙️ Margin Overrides for {d['ticker']}**")
    st.caption("Auto-filled from real financials — adjust if needed")

    ebit_margin = st.slider("EBIT Margin %",        1.0, 50.0,
                            max(1.0, round(d["ebit_margin"]*100, 1)), 0.5) / 100
    tax_rate    = st.slider("Tax Rate %",           5.0, 40.0,
                            round(d["tax_rate"]*100, 0), 1.0) / 100
    da_pct      = st.slider("D&A % of Revenue",    0.5, 15.0,
                            max(0.5, round(d["da_pct"]*100, 1)), 0.5) / 100
    capex_pct   = st.slider("CapEx % of Revenue",  0.5, 20.0,
                            max(0.5, round(d["capex_pct"]*100, 1)), 0.5) / 100
    nwc_pct     = st.slider("NWC Change % of Rev", 0.0, 5.0, 0.5, 0.25) / 100

    st.divider()
    st.markdown("**🏢 Balance Sheet Overrides**")
    gross_debt  = st.number_input("Gross Debt ($mn)",   value=max(0, d["debt"]),   step=500)
    cash_val    = st.number_input("Cash ($mn)",         value=max(0, d["cash"]),   step=500)
    non_op      = st.number_input("Non-Op Assets ($mn)",value=max(0, d["non_op"]), step=100)
    nci         = st.number_input("NCI ($mn)",          value=0,                   step=10)
    shares_out  = st.number_input("Diluted Shares (mn)",value=max(1.0, d["shares"]), step=10.0)

equity_wt = 1 - debt_wt

# ─────────────────────────────────────────────────────────────────
# DCF ENGINE
# ─────────────────────────────────────────────────────────────────

@st.cache_data
def run_dcf(base_rev, rev_growth, ebit_mgn, tax_rt, da_p, capex_p, nwc_p,
            rf, beta, erp, add_rp, pretax_kd, debt_wt, equity_wt,
            ltgr, exit_mult, gross_debt, cash_val, non_op, nci, shares_out):

    # ── Projections ──────────────────────────────────────────────
    revenues, ebits, taxes, nopats = [], [], [], []
    das, ebitdas, capexs, nwcs, fcffs = [], [], [], [], []

    for i in range(5):
        rev   = (revenues[-1] if i else base_rev) * (1 + rev_growth[i])
        ebit  = rev * ebit_mgn
        tax   = ebit * tax_rt
        nopat = ebit - tax
        da    = rev * da_p
        capex = rev * capex_p
        nwc   = rev * nwc_p
        fcff  = nopat + da - nwc - capex
        revenues.append(round(rev)); ebits.append(round(ebit))
        taxes.append(round(tax));    nopats.append(round(nopat))
        das.append(round(da));       ebitdas.append(round(ebit+da))
        capexs.append(round(capex)); nwcs.append(round(nwc))
        fcffs.append(round(fcff))

    # ── WACC ─────────────────────────────────────────────────────
    ke   = rf + beta * erp + add_rp
    kd   = pretax_kd * (1 - tax_rt)
    wacc = equity_wt * ke + debt_wt * kd

    # ── Discount (mid-year) ───────────────────────────────────────
    pv_f    = [(1/(1+wacc)**(i+0.5)) for i in range(5)]
    pv_fcff = [round(f*p) for f, p in zip(fcffs, pv_f)]
    pv_exp  = sum(pv_fcff)

    # ── Terminal Value ────────────────────────────────────────────
    tv_gg    = fcffs[-1]*(1+ltgr)/(wacc-ltgr) if wacc > ltgr else 0
    pv_tv_gg = round(tv_gg * pv_f[-1])
    ev_gg    = pv_exp + pv_tv_gg
    eq_gg    = ev_gg + non_op + cash_val - gross_debt - nci
    pps_gg   = round(eq_gg / shares_out, 2) if shares_out else 0

    tv_exit  = ebitdas[-1] * exit_mult
    pv_tv_ex = round(tv_exit * pv_f[-1])
    ev_exit  = pv_exp + pv_tv_ex
    eq_exit  = ev_exit + non_op + cash_val - gross_debt - nci
    pps_exit = round(eq_exit / shares_out, 2) if shares_out else 0

    return dict(
        revenues=revenues, ebits=ebits, taxes=taxes, nopats=nopats,
        das=das, ebitdas=ebitdas, capexs=capexs, nwcs=nwcs, fcffs=fcffs,
        pv_f=[round(p,4) for p in pv_f], pv_fcff=pv_fcff, pv_exp=round(pv_exp),
        ke=round(ke,4), kd=round(kd,4), wacc=round(wacc,4), beta=beta,
        tv_gg=round(tv_gg),   pv_tv_gg=pv_tv_gg,
        ev_gg=round(ev_gg),   eq_gg=round(eq_gg),   pps_gg=pps_gg,
        tv_exit=round(tv_exit), pv_tv_ex=pv_tv_ex,
        ev_exit=round(ev_exit), eq_exit=round(eq_exit), pps_exit=pps_exit,
    )

m = run_dcf(
    d["revenue"], tuple(revenue_growth), ebit_margin, tax_rate,
    da_pct, capex_pct, nwc_pct,
    rf_rate, d["beta"], erp, add_rp, pretax_kd, debt_wt, equity_wt,
    ltgr, exit_mult, gross_debt, cash_val, non_op, nci, shares_out
)

# ─────────────────────────────────────────────────────────────────
# MAIN PAGE
# ─────────────────────────────────────────────────────────────────

# Header
upside_gg   = ((m['pps_gg']   / d['price'] - 1)*100) if d['price'] else 0
upside_exit = ((m['pps_exit'] / d['price'] - 1)*100) if d['price'] else 0
avg_pps     = (m['pps_gg'] + m['pps_exit']) / 2

st.markdown(f"""
<div class="dcf-header">
    <div style="display:flex; align-items:center; gap:12px; margin-bottom:6px">
        <span class="company-pill">{d['ticker']}</span>
        <span style="color:#aac;font-size:13px">{d['exchange']}</span>
    </div>
    <div class="dcf-title">{d['name']}</div>
    <div class="dcf-sub">
        {d['sector']} · {d['industry']} &nbsp;|&nbsp;
        Current Price: <b style="color:#fff">${d['price']:,.2f}</b> &nbsp;|&nbsp;
        Market Cap: <b style="color:#fff">${d['mktcap']:,}mn</b> &nbsp;|&nbsp;
        WACC: <b style="color:#f59e0b">{m['wacc']*100:.2f}%</b>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Result cards ─────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
with c1:
    clr = "#00c878" if upside_gg >= 0 else "#ff4b4b"
    st.markdown(f"""<div class="result-card">
        <div class="result-label">Fair Value — Gordon Growth</div>
        <div class="result-value" style="color:{clr}">${m['pps_gg']:.2f}</div>
        <div class="result-sub">{'▲' if upside_gg>=0 else '▼'} {abs(upside_gg):.1f}% vs current price</div>
    </div>""", unsafe_allow_html=True)
with c2:
    clr2 = "#6001d2" if upside_exit >= 0 else "#ff4b4b"
    st.markdown(f"""<div class="result-card">
        <div class="result-label">Fair Value — Exit Multiple</div>
        <div class="result-value" style="color:{clr2}">${m['pps_exit']:.2f}</div>
        <div class="result-sub">{'▲' if upside_exit>=0 else '▼'} {abs(upside_exit):.1f}% vs current price</div>
    </div>""", unsafe_allow_html=True)
with c3:
    clr3 = "#f59e0b" if (avg_pps/d['price']-1) >= 0 else "#ff4b4b"
    st.markdown(f"""<div class="result-card">
        <div class="result-label">Blended Fair Value</div>
        <div class="result-value" style="color:{clr3}">${avg_pps:.2f}</div>
        <div class="result-sub">Average of both methods</div>
    </div>""", unsafe_allow_html=True)
with c4:
    st.markdown(f"""<div class="result-card">
        <div class="result-label">WACC</div>
        <div class="result-value">{m['wacc']*100:.2f}%</div>
        <div class="result-sub">Ke {m['ke']*100:.2f}% · Kd {m['kd']*100:.2f}%</div>
    </div>""", unsafe_allow_html=True)

# ── Upside / downside indicator ──────────────────────────────────
verdict_pps  = avg_pps
if d['price'] and verdict_pps:
    upside = (verdict_pps/d['price'] - 1)*100
    if upside > 15:
        verdict_color, verdict_label = "#00c878", f"📈 UNDERVALUED — {upside:.1f}% upside potential"
    elif upside < -15:
        verdict_color, verdict_label = "#ff4b4b", f"📉 OVERVALUED — {abs(upside):.1f}% downside risk"
    else:
        verdict_color, verdict_label = "#f59e0b", f"⚖️ FAIRLY VALUED — {upside:+.1f}% vs current"
    st.markdown(f"""
    <div style="background:{verdict_color}22; border:1px solid {verdict_color}55;
                border-radius:10px; padding:12px 20px; margin:12px 0;
                color:{verdict_color}; font-weight:700; font-size:14px; text-align:center">
        {verdict_label} &nbsp;·&nbsp; Current: ${d['price']:.2f} &nbsp;·&nbsp; Blended DCF: ${verdict_pps:.2f}
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── TABS ─────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📈 Projections", "💵 DCF Bridge", "🏦 WACC", "🎯 Sensitivity", "📥 Export"
])

# ──────────────────────────────────────────────────────────────────
# TAB 1 — PROJECTIONS
# ──────────────────────────────────────────────────────────────────
with tab1:
    col1, col2 = st.columns(2)
    with col1:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=YEARS, y=m['revenues'], name='Revenue', marker_color='#1B3A6B', opacity=0.85))
        fig.add_trace(go.Bar(x=YEARS, y=m['ebits'],    name='EBIT',    marker_color='#6001d2', opacity=0.85))
        fig.update_layout(barmode='group', height=320, **DARK, title="Revenue vs EBIT ($mn)")
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(x=YEARS, y=m['ebitdas'], name='EBITDA',  marker_color='#2d5aa0', opacity=0.85))
        fig2.add_trace(go.Scatter(x=YEARS, y=m['fcffs'], name='FCFF',
                                  line=dict(color='#00c878', width=3),
                                  mode='lines+markers', marker=dict(size=8)))
        fig2.update_layout(height=320, **DARK, title="EBITDA vs FCFF ($mn)")
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown('<div class="section-title">Projected Income Statement ($mn)</div>', unsafe_allow_html=True)

    inc_rows = [
        ("Revenue",     [f"${v:,}" for v in m['revenues']], False),
        ("EBIT",        [f"${v:,}" for v in m['ebits']],    False),
        ("EBIT Margin", [f"{v/r*100:.1f}%" for v,r in zip(m['ebits'], m['revenues'])], True),
        ("Tax",         [f"${v:,}" for v in m['taxes']],    False),
        ("NOPAT",       [f"${v:,}" for v in m['nopats']],   False),
        ("D&A",         [f"${v:,}" for v in m['das']],      False),
        ("EBITDA",      [f"${v:,}" for v in m['ebitdas']],  False),
        ("CapEx",       [f"(${v:,})" for v in m['capexs']], False),
        ("FCFF",        [f"${v:,}" for v in m['fcffs']],    False),
    ]

    year_headers = "".join(f'<th style="background:#1B3A6B;color:#fff;padding:10px 16px;text-align:center;font-size:13px;font-weight:700;border:1px solid #2d5aa0">{y}</th>' for y in YEARS)
    table_rows = ""
    for i, (label, vals, is_pct) in enumerate(inc_rows):
        bg       = "#dbeafe" if i % 2 == 0 else "#ffffff"
        lbl_bg   = "#1B3A6B" if label in ("FCFF", "EBITDA") else "#1e4080"
        lbl_fw   = "800" if label in ("FCFF", "EBITDA") else "600"
        val_fw   = "700" if label in ("FCFF", "EBITDA") else "500"
        val_col  = "#1B3A6B" if not is_pct else "#0f4c9e"
        cells    = "".join(f'<td style="padding:9px 16px;text-align:right;font-size:13px;background:{bg};color:{val_col};font-weight:{val_fw};border:1px solid #bfdbfe">{v}</td>' for v in vals)
        table_rows += f'<tr><td style="padding:9px 14px;font-size:13px;font-weight:{lbl_fw};background:{lbl_bg};color:#fff;border:1px solid #2d5aa0;white-space:nowrap">{label}</td>{cells}</tr>'

    st.markdown(f"""
    <div style="overflow-x:auto; border-radius:10px; border:1px solid #1B3A6B; margin-top:8px">
    <table style="width:100%;border-collapse:collapse;font-family:Inter,sans-serif">
      <thead>
        <tr>
          <th style="background:#0f2a5c;color:#fff;padding:10px 14px;text-align:left;font-size:13px;font-weight:700;border:1px solid #2d5aa0">Metric</th>
          {year_headers}
        </tr>
      </thead>
      <tbody>{table_rows}</tbody>
    </table>
    </div>
    """, unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────
# TAB 2 — DCF BRIDGE
# ──────────────────────────────────────────────────────────────────
with tab2:
    col1, col2 = st.columns(2)
    with col1:
        labels = [str(y) for y in YEARS] + ["Terminal\nValue (GG)"]
        values = m['pv_fcff'] + [m['pv_tv_gg']]
        fig3 = go.Figure(go.Bar(x=labels, y=values,
                                marker_color=['#1B3A6B']*5+['#00c878'], opacity=0.9))
        fig3.update_layout(height=300, **DARK, title="PV of Cash Flows — Gordon Growth ($mn)")
        st.plotly_chart(fig3, use_container_width=True)
    with col2:
        fig4 = go.Figure(go.Waterfall(
            orientation="v",
            measure=["relative","relative","total","relative","relative","relative","total"],
            x=["PV FCFs","PV TV (GG)","Enterprise\nValue","+ Cash","+ Non-Op","- Debt","Equity\nValue"],
            y=[m['pv_exp'], m['pv_tv_gg'], 0, cash_val, non_op, -gross_debt, 0],
            connector=dict(line=dict(color="#444")),
            increasing=dict(marker=dict(color="#00c878")),
            decreasing=dict(marker=dict(color="#ff4b4b")),
            totals=dict(marker=dict(color="#1B3A6B")),
        ))
        fig4.update_layout(height=300, **DARK, title="Equity Value Waterfall ($mn)")
        st.plotly_chart(fig4, use_container_width=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown('<div class="section-title">Method 1: Gordon Growth</div>', unsafe_allow_html=True)
        for label, val in [
            ("PV Explicit FCFs",    f"${m['pv_exp']:,}mn"),
            ("Terminal Value",      f"${m['tv_gg']:,}mn"),
            ("PV of TV",            f"${m['pv_tv_gg']:,}mn"),
            ("Enterprise Value",    f"${m['ev_gg']:,}mn"),
            ("Equity Value",        f"${m['eq_gg']:,}mn"),
            ("★ Price per Share",   f"${m['pps_gg']:.2f}"),
        ]:
            st.metric(label, val)
    with col_b:
        st.markdown('<div class="section-title">Method 2: Exit Multiple</div>', unsafe_allow_html=True)
        for label, val in [
            ("PV Explicit FCFs",          f"${m['pv_exp']:,}mn"),
            (f"TV ({exit_mult}x EBITDA)", f"${m['tv_exit']:,}mn"),
            ("PV of TV",                  f"${m['pv_tv_ex']:,}mn"),
            ("Enterprise Value",          f"${m['ev_exit']:,}mn"),
            ("Equity Value",              f"${m['eq_exit']:,}mn"),
            ("★ Price per Share",         f"${m['pps_exit']:.2f}"),
        ]:
            st.metric(label, val)

# ──────────────────────────────────────────────────────────────────
# TAB 3 — WACC
# ──────────────────────────────────────────────────────────────────
with tab3:
    col1, col2, col3 = st.columns(3)
    wacc_items = [
        (col1, "Cost of Equity (CAPM)", [
            ("Risk-Free Rate",        f"{rf_rate*100:.2f}%"),
            ("Beta (β)",              f"{d['beta']:.2f}"),
            ("Equity Risk Premium",   f"{erp*100:.2f}%"),
            ("Additional Premium",    f"{add_rp*100:.2f}%"),
            ("= Cost of Equity",      f"{m['ke']*100:.2f}%", "#00c878"),
        ]),
        (col2, "Cost of Debt", [
            ("Pre-Tax Cost of Debt",  f"{pretax_kd*100:.2f}%"),
            ("Tax Rate",              f"{tax_rate*100:.0f}%"),
            ("= After-Tax Cost Debt", f"{m['kd']*100:.2f}%", "#6001d2"),
        ]),
        (col3, "WACC", [
            ("Debt Weight",           f"{debt_wt*100:.0f}%"),
            ("Equity Weight",         f"{equity_wt*100:.0f}%"),
            ("= WACC",                f"{m['wacc']*100:.2f}%", "#f59e0b"),
        ]),
    ]
    for col, title, items in wacc_items:
        with col:
            st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)
            for item in items:
                border = f"border-color:{item[2]};" if len(item)>2 else ""
                color  = f"color:{item[2]};" if len(item)>2 else ""
                st.markdown(f"""<div class="metric-card" style="{border}">
                    <div class="metric-label">{item[0]}</div>
                    <div class="metric-value" style="{color}">{item[1]}</div>
                </div>""", unsafe_allow_html=True)

    fig_w = go.Figure(go.Pie(
        labels=['Debt', 'Equity'],
        values=[debt_wt*100, equity_wt*100],
        hole=0.55,
        marker_colors=['#6001d2', '#1B3A6B'],
    ))
    fig_w.update_layout(height=250, **DARK, title="Capital Structure")
    st.plotly_chart(fig_w, use_container_width=True)

# ──────────────────────────────────────────────────────────────────
# TAB 4 — SENSITIVITY
# ──────────────────────────────────────────────────────────────────
with tab4:
    st.markdown('<div class="section-title">Sensitivity: WACC vs Long-Term Growth Rate → Implied Share Price ($)</div>', unsafe_allow_html=True)

    wacc_range = [round(m['wacc'] + d2, 4) for d2 in [-0.015,-0.010,-0.005,0,0.005,0.010,0.015]]
    ltgr_range = [0.015, 0.020, 0.025, 0.030, 0.035]
    base_pps   = m['pps_gg']

    rows = {}
    for w in wacc_range:
        row = {}
        pv_f2  = [(1/(1+w)**(i+0.5)) for i in range(5)]
        pv_ex2 = sum(f*p for f,p in zip(m['fcffs'], pv_f2))
        for g in ltgr_range:
            col_name = f"{g*100:.1f}%"
            if w <= g:
                row[col_name] = None
            else:
                tv2  = m['fcffs'][-1]*(1+g)/(w-g)
                pv2  = tv2 * pv_f2[-1]
                eq2  = pv_ex2 + pv2 + non_op + cash_val - gross_debt - nci
                row[col_name] = round(eq2/shares_out, 2)
        rows[f"{w*100:.1f}%"] = row

    df_sens = pd.DataFrame(rows).T
    df_sens.index.name = "WACC \\ LTGR"

    def color_cell(val):
        if val is None: return 'background-color:#0f1c30; color:#3a5a80; font-style:italic'
        if val > base_pps*1.15: return 'background-color:#1B3A6B; color:#ffffff; font-weight:bold'
        if val < base_pps*0.85: return 'background-color:#4a7ab5; color:#ffffff; font-weight:bold'
        return 'background-color:#2d5aa0; color:#ffffff'

    styled = df_sens.style.map(color_cell).format(
        lambda x: f"${x:.2f}" if x is not None else "N/A"
    )
    st.dataframe(styled, use_container_width=True, height=300)
    st.caption("🔵 Dark Navy = >15% above base  ·  🔷 Mid Blue = base range  ·  🩵 Light Blue = >15% below base")

    # Tornado
    st.markdown('<div class="section-title">Key Driver Impact on Share Price</div>', unsafe_allow_html=True)
    def quick_pps(rev_g=None, ebit_m=None, w=None, g=None, cp=None):
        rg = rev_g or tuple(revenue_growth)
        em = ebit_m or ebit_margin
        wc = w or m['wacc']
        gc = g or ltgr
        cc = cp or capex_pct
        fcffs2 = []
        revs2  = []
        for i in range(5):
            rv = (revs2[-1] if i else d["revenue"]) * (1 + rg[i])
            fc = rv*em*(1-tax_rate) + rv*da_pct - rv*nwc_pct - rv*cc
            revs2.append(rv); fcffs2.append(fc)
        pf2 = [(1/(1+wc)**(i+0.5)) for i in range(5)]
        pe2 = sum(f*p for f,p in zip(fcffs2,pf2))
        tv2 = fcffs2[-1]*(1+gc)/(wc-gc) if wc>gc else 0
        eq2 = pe2 + tv2*pf2[-1] + non_op + cash_val - gross_debt - nci
        return round(eq2/shares_out, 2)

    impacts = {
        "Revenue Growth +2%":  quick_pps(rev_g=tuple([g+0.02 for g in revenue_growth])) - base_pps,
        "EBIT Margin +2%":     quick_pps(ebit_m=ebit_margin+0.02) - base_pps,
        "WACC -0.5%":          quick_pps(w=m['wacc']-0.005) - base_pps,
        "LTGR +0.5%":          quick_pps(g=ltgr+0.005) - base_pps,
        "CapEx -2% Revenue":   quick_pps(cp=capex_pct-0.02) - base_pps,
    }
    fig_t = go.Figure(go.Bar(
        x=list(impacts.values()), y=list(impacts.keys()), orientation='h',
        marker_color=['#1B3A6B' if v>=0 else '#2d5aa0' for v in impacts.values()],
        marker_line=dict(color='#aac', width=0.5),
        opacity=0.95,
        text=[f"+${v:.2f}" if v>=0 else f"-${abs(v):.2f}" for v in impacts.values()],
        textposition='outside',
        textfont=dict(color='#aac', size=11),
    ))
    fig_t.add_vline(x=0, line_dash="dash", line_color="#aac", line_width=1.5)
    fig_t.update_layout(height=280, **DARK, title=f"Key Driver Impact on Share Price ($) — Base: ${base_pps:.2f}")
    st.plotly_chart(fig_t, use_container_width=True)

# ──────────────────────────────────────────────────────────────────
# TAB 5 — EXPORT
# ──────────────────────────────────────────────────────────────────
with tab5:
    st.markdown('<div class="section-title">Download Excel Report</div>', unsafe_allow_html=True)

    def build_excel():
        wb = Workbook()

        # ── Styles ────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="1B3A6B")
        sub_fill  = PatternFill("solid", fgColor="2d5aa0")
        alt_fill  = PatternFill("solid", fgColor="dbeafe")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        sub_font  = Font(bold=True, color="FFFFFF", size=10)
        ctr       = Alignment(horizontal="center", vertical="center")
        rgt       = Alignment(horizontal="right",  vertical="center")

        def hdr(ws, row, col, val, fill=hdr_fill, font=hdr_font, align=ctr):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.font = font; c.alignment = align

        def val(ws, row, col, v, fill=None, font=None, align=rgt):
            c = ws.cell(row=row, column=col, value=v)
            if fill: c.fill = fill
            if font: c.font = font
            c.alignment = align

        # ── Sheet 1: Summary ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "DCF Summary"
        ws1.column_dimensions["A"].width = 28
        for col in ["B","C","D","E","F"]: ws1.column_dimensions[col].width = 16

        hdr(ws1, 1, 1, f"DCF Valuation — {d['name']} ({d['ticker']})", fill=hdr_fill, font=Font(bold=True,color="FFFFFF",size=13), align=Alignment(horizontal="left"))
        ws1.merge_cells("A1:F1")
        hdr(ws1, 2, 1, f"Sector: {d['sector']}  |  WACC: {m['wacc']*100:.2f}%  |  Current Price: ${d['price']:.2f}", fill=sub_fill, font=sub_font, align=Alignment(horizontal="left"))
        ws1.merge_cells("A2:F2")

        results_data = [
            ("Gordon Growth — Price/Share", f"${m['pps_gg']:.2f}"),
            ("Exit Multiple — Price/Share",  f"${m['pps_exit']:.2f}"),
            ("Blended Fair Value",            f"${(m['pps_gg']+m['pps_exit'])/2:.2f}"),
            ("Current Market Price",          f"${d['price']:.2f}"),
            ("Enterprise Value (GG)",         f"${m['ev_gg']:,}mn"),
            ("Enterprise Value (Exit)",       f"${m['ev_exit']:,}mn"),
            ("WACC",                          f"{m['wacc']*100:.2f}%"),
            ("Cost of Equity",                f"{m['ke']*100:.2f}%"),
            ("After-Tax Cost of Debt",        f"{m['kd']*100:.2f}%"),
        ]
        hdr(ws1, 4, 1, "Valuation Output", fill=hdr_fill)
        hdr(ws1, 4, 2, "Value",             fill=hdr_fill)
        for i, (lbl, v) in enumerate(results_data):
            f = alt_fill if i % 2 == 0 else None
            val(ws1, 5+i, 1, lbl, fill=f, font=Font(size=10), align=Alignment(horizontal="left"))
            val(ws1, 5+i, 2, v,   fill=f, font=Font(bold=True, color="1B3A6B", size=10))

        # ── Sheet 2: Income Statement ─────────────────────────────
        ws2 = wb.create_sheet("Income Statement")
        ws2.column_dimensions["A"].width = 22
        for col in ["B","C","D","E","F"]: ws2.column_dimensions[col].width = 16

        hdr(ws2, 1, 1, "Projected Income Statement ($mn)", fill=hdr_fill, font=Font(bold=True,color="FFFFFF",size=13), align=Alignment(horizontal="left"))
        ws2.merge_cells("A1:F1")
        hdr(ws2, 2, 1, "Metric", fill=sub_fill, font=sub_font, align=Alignment(horizontal="left"))
        for j, y in enumerate(YEARS):
            hdr(ws2, 2, j+2, str(y), fill=sub_fill, font=sub_font)

        is_rows = [
            ("Revenue",     m['revenues']),
            ("EBIT",        m['ebits']),
            ("EBIT Margin", [f"{v/r*100:.1f}%" for v,r in zip(m['ebits'],m['revenues'])]),
            ("Tax",         m['taxes']),
            ("NOPAT",       m['nopats']),
            ("D&A",         m['das']),
            ("EBITDA",      m['ebitdas']),
            ("CapEx",       [f"({v:,})" for v in m['capexs']]),
            ("FCFF",        m['fcffs']),
        ]
        for i, (lbl, vals) in enumerate(is_rows):
            f = alt_fill if i % 2 == 0 else None
            bold_row = lbl in ("FCFF", "EBITDA", "Revenue")
            val(ws2, 3+i, 1, lbl, fill=f, font=Font(bold=bold_row, size=10), align=Alignment(horizontal="left"))
            for j, v in enumerate(vals):
                disp = f"${v:,}" if isinstance(v, int) else v
                val(ws2, 3+i, j+2, disp, fill=f, font=Font(bold=bold_row, color="1B3A6B", size=10))

        # ── Sheet 3: DCF Bridge ────────────────────────────────────
        ws3 = wb.create_sheet("DCF Bridge")
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 18
        hdr(ws3, 1, 1, "DCF Bridge ($mn)", fill=hdr_fill, font=Font(bold=True,color="FFFFFF",size=13), align=Alignment(horizontal="left"))
        ws3.merge_cells("A1:B1")

        def fmt_val(v):
            if isinstance(v, float):
                return f"$({abs(v):,.2f})" if v < 0 else f"${v:,.2f}"
            if isinstance(v, int):
                return f"$({abs(v):,})" if v < 0 else f"${v:,}"
            return str(v)

        bridge_rows_gg = [
            ("PV of Explicit FCFs",       m['pv_exp']),
            ("PV of Terminal Value (GG)", m['pv_tv_gg']),
            ("Enterprise Value",          m['ev_gg']),
            ("Add: Cash",                 cash_val),
            ("Add: Non-Operating Assets", non_op),
            ("Less: Gross Debt",          -gross_debt),
            ("Less: NCI",                 -nci),
            ("Equity Value",              m['eq_gg']),
            ("Diluted Shares (mn)",       shares_out),
            ("Price Per Share (GG)",      m['pps_gg']),
        ]
        hdr(ws3, 2, 1, "Gordon Growth Method", fill=sub_fill, font=sub_font, align=Alignment(horizontal="left"))
        hdr(ws3, 2, 2, "Value ($mn)", fill=sub_fill, font=sub_font)
        for i, (lbl, v) in enumerate(bridge_rows_gg):
            f = alt_fill if i % 2 == 0 else None
            bold_row = lbl in ("Equity Value", "Enterprise Value", "Price Per Share (GG)")
            val(ws3, 3+i, 1, lbl, fill=f, font=Font(bold=bold_row, size=10), align=Alignment(horizontal="left"))
            val(ws3, 3+i, 2, fmt_val(v), fill=f, font=Font(bold=bold_row, color="1B3A6B", size=10))

        # ── Sheet 4: WACC ─────────────────────────────────────────
        ws4 = wb.create_sheet("WACC")
        ws4.column_dimensions["A"].width = 28
        ws4.column_dimensions["B"].width = 16
        hdr(ws4, 1, 1, "WACC Calculation", fill=hdr_fill, font=Font(bold=True,color="FFFFFF",size=13), align=Alignment(horizontal="left"))
        ws4.merge_cells("A1:B1")
        wacc_data = [
            ("Risk-Free Rate",            f"{rf_rate*100:.2f}%"),
            ("Beta", f"{d['beta']:.2f}"),
            ("Equity Risk Premium",       f"{erp*100:.2f}%"),
            ("Additional Risk Premium",   f"{add_rp*100:.2f}%"),
            ("Cost of Equity (Ke)",       f"{m['ke']*100:.2f}%"),
            ("Pre-Tax Cost of Debt",      f"{pretax_kd*100:.2f}%"),
            ("Tax Rate",                  f"{tax_rate*100:.0f}%"),
            ("After-Tax Cost of Debt",    f"{m['kd']*100:.2f}%"),
            ("Debt Weight",               f"{debt_wt*100:.0f}%"),
            ("Equity Weight",             f"{equity_wt*100:.0f}%"),
            ("WACC",                      f"{m['wacc']*100:.2f}%"),
        ]
        for i, (lbl, v) in enumerate(wacc_data):
            f = alt_fill if i % 2 == 0 else None
            bold_row = lbl == "WACC"
            val(ws4, 2+i, 1, lbl, fill=f, font=Font(bold=bold_row, size=10), align=Alignment(horizontal="left"))
            val(ws4, 2+i, 2, v,   fill=f, font=Font(bold=bold_row, color="1B3A6B", size=10))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    st.markdown(f"""
    <div style="background:#1a2a3a; border:1px solid #1B3A6B; border-radius:10px; padding:20px 24px; margin-bottom:16px">
        <div style="font-size:15px; font-weight:700; color:#fff; margin-bottom:6px">📊 {d['name']} — DCF Valuation Report</div>
        <div style="font-size:12px; color:#aac">Includes: DCF Summary · Projected Income Statement · DCF Bridge · WACC Breakdown</div>
    </div>
    """, unsafe_allow_html=True)

    excel_bytes = build_excel()
    filename    = f"{d['ticker']}_DCF_Valuation_{datetime.today().strftime('%Y%m%d')}.xlsx"

    st.download_button(
        label="Download Excel Report",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
